import openpyxl
import json
import sys

try:
    wb = openpyxl.load_workbook('灯　車両日報.xlsx', data_only=True)
    ws = wb.active

    data = {
        'max_row': ws.max_row,
        'max_column': ws.max_column,
        'merged_cells': [str(merge) for merge in ws.merged_cells.ranges],
        'row_dimensions': {},
        'col_dimensions': {},
        'cells': []
    }

    # Extract some row heights and col widths
    for r in range(1, min(ws.max_row + 1, 50)):
        if ws.row_dimensions[r].height is not None:
            data['row_dimensions'][r] = ws.row_dimensions[r].height

    for c in range(1, min(ws.max_column + 1, 20)):
        col_letter = openpyxl.utils.get_column_letter(c)
        if ws.column_dimensions[col_letter].width is not None:
            data['col_dimensions'][col_letter] = ws.column_dimensions[col_letter].width

    # Extract non-empty cells
    for r in range(1, min(ws.max_row + 1, 50)):
        for c in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                cell_info = {
                    'coord': cell.coordinate,
                    'val': cell.value,
                    'align': {
                        'hoz': cell.alignment.horizontal if cell.alignment else None,
                        'vert': cell.alignment.vertical if cell.alignment else None,
                    },
                    'border': {
                        'bottom': cell.border.bottom.style if cell.border.bottom else None,
                        'top': cell.border.top.style if cell.border.top else None,
                        'left': cell.border.left.style if cell.border.left else None,
                        'right': cell.border.right.style if cell.border.right else None,
                    }
                }
                data['cells'].append(cell_info)

    print(json.dumps(data, indent=2, ensure_ascii=False))

except Exception as e:
    print(str(e), file=sys.stderr)
